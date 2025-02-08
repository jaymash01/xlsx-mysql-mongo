from openpyxl.reader.excel import load_workbook
from datetime import datetime
from mysql_connection import get_connection

connection = get_connection()

_path = "files/{}"


def import_draws():
    print("Importing draws...")
    sheet = load_workbook(_path.format("draw.xlsx")).active
    rows = sheet.iter_rows(values_only=True, min_row=2)
    cursor = connection.cursor()

    for row in rows:
        print(row)
        created_date = datetime.strptime(row[1], '%d-%b-%y %I.%M.%S.%f %p')  # .astimezone(tz=timezone.utc)
        sql = "insert ignore into draw(id, createdDate, numberOfTheDay, isActive) values (%s, %s, %s, %s)"
        values = (row[0], created_date, row[3], row[2])

        cursor.execute(sql, values)

    cursor.close()
    connection.commit()
    print("Imported draws successfully...")


def import_players():
    print("Importing players...")

    sheet = load_workbook(_path.format("player.xlsx")).active
    rows = sheet.iter_rows(values_only=True, min_row=2)
    cursor = connection.cursor()

    for row in rows:
        print(row)
        created_date = datetime.strptime(row[1], '%d-%b-%y %I.%M.%S.%f %p')
        last_played_date = datetime.strptime(row[2], '%d-%b-%y %I.%M.%S.%f %p') if row[2] is not None else None
        sql = "insert ignore into player(id, msisdn, operator, createdDate, lastPlayedDate) values (%s, %s, %s, %s, %s)"
        values = (row[0], row[3], row[4], created_date, last_played_date)

        cursor.execute(sql, values)

    cursor.close()
    connection.commit()
    print("Imported players successfully...")


def import_collections():
    print("Importing collections...")

    sheet = load_workbook(_path.format("collection.xlsx")).active
    rows = sheet.iter_rows(values_only=True, min_row=2)
    cursor = connection.cursor()

    for row in rows:
        print(row)
        created_date = datetime.strptime(row[5], '%d-%b-%y %I.%M.%S.%f %p')
        sql = "insert ignore into mobileMoney(id, createdDate, company, operator, txnId, paybillNumber, receipt, msisdn, amount, reference, country, numberChoice, statusId, transactionTypeId, playerId, processed, description) values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        values = (
            row[0], created_date, row[3], row[7], row[12], row[8], row[10], row[6],
            row[2], row[11], row[4], row[1], row[14], row[15], row[13], row[9], row[16],
        )

        cursor.execute(sql, values)

    cursor.close()
    connection.commit()
    print("Imported collections successfully...")


def import_disbursements():
    print("Importing disbursements...")

    sheet = load_workbook(_path.format("disbursement.xlsx")).active
    rows = sheet.iter_rows(values_only=True, min_row=2)
    cursor = connection.cursor()

    for row in rows:
        print(row)
        created_date = datetime.strptime(row[5], '%d-%b-%y %I.%M.%S.%f %p')
        sql = "insert ignore into mobileMoney(id, createdDate, company, operator, txnId, paybillNumber, receipt, msisdn, amount, reference, country, numberChoice, statusId, transactionTypeId, playerId, processed, description) values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        values = (
            row[0], created_date, row[3], row[7], row[12], row[8], row[10], row[6],
            row[2], row[11], row[4], row[1], row[14], row[15], row[13], row[9], row[16],
        )

        cursor.execute(sql, values)

    cursor.close()
    connection.commit()
    print("Imported disbursements successfully...")


def import_transactions():
    print("Importing transactions...")

    sheet = load_workbook(_path.format("transaction.xlsx")).active
    rows = sheet.iter_rows(values_only=True, min_row=2)
    cursor = connection.cursor()

    for row in rows:
        print(row)
        created_date = datetime.strptime(row[2], '%d-%b-%y %I.%M.%S.%f %p')
        updated_date = datetime.strptime(row[8], '%d-%b-%y %I.%M.%S.%f %p') if row[8] is not None else None
        sql = "insert ignore into transaction(id, ticket, numberOfTheDay, jackpotNumber, createdDate, updatedDate, playerId, mobileMoneyId, drawId, isPaid, numberChoice, winningAmount, payoutAmount, luckNumberId, createdAt) values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        values = (
            row[0], row[7], row[5], row[4], created_date, updated_date, row[12], row[11],
            row[10], row[3], row[1], row[9], row[6], row[14], created_date,
        )

        cursor.execute(sql, values)

    cursor.close()
    connection.commit()
    print("Imported transactions successfully...")
